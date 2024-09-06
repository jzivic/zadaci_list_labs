from A_prep import *

import openpyxl
from sklearn.svm import SVC






gamma_range = range(2)
C_range = range(2)

C_f = lambda n: 2 ** n
gamma_f = lambda n: 2 ** n


xlsx_name = "output/Acc_SVM.xlsx"  # xlsx file path for saving results
Acc_grid_search = pd.ExcelWriter(xlsx_name)  # creates xlsx file








class Model_1(DataOverview):

    def __init__(self, csv_data):
        super().__init__(csv_data)

        self.svm_model()





    def calc_all_acc(self, y_true, y_prediction):
        tn, fp, fn, tp = confusion_matrix(y_true, y_prediction).ravel()
        tpr = round(tp / (tp + fn), 3) * 100
        fpr = round(fp / (fp + tn), 3) * 100
        fnr = round(fn / (tp + fn), 3) * 100
        tnr = round(tn / (fp + tn), 3) * 100
        return {"tpr": tpr, "fpr": fpr, "fnr": fnr, "tnr": tnr}





    def svm_model(self):

        best_grid_acc = []  # index coordinates for best acc in grid search
        accuracity_matrix = {}  # matrix for saving acc
        best_acc = {"fpr": 100, "tpr": 0}  # acc initiation

        for gamma in gamma_range:  # searching for gamma with one same C value set

            print(gamma)

            gamma_str = "{:.2e}".format(gamma_f(gamma))  # gamma string for xlsx file
            accuracity_matrix[gamma_str] = []  # acc list for gamma values


            for c in C_range:                              # searching for C with one same gamma value set
                svcModel = SVC(C=C_f(c), kernel="rbf", gamma=gamma_f(gamma))
                svcModel.fit(self.X_train, self.y_train)
                y_pred = svcModel.predict(self.X_test)
                all_acc = self.calc_all_acc(self.y_test, y_pred)
                tpr, fpr, fnr, tnr = all_acc["tpr"], all_acc["fpr"], all_acc["fnr"], all_acc["tnr"]

                print(tpr, fpr, fnr, tnr )


                if fpr < best_acc["fpr"]:
                    n_gamma, n_C = gamma_range.index(gamma), C_range.index(c)   # index of best gamma and C value
                    best_acc["fpr"],  best_acc["tpr"] = fpr, tpr
                    best_grid_acc = [[n_C, n_gamma]]

                # in case of same fpr, model will select parameters with better tpr acc
                if fpr == best_acc["fpr"] and tpr > best_acc["tpr"]:
                    best_acc["tpr"] = tpr
                    best_grid_acc = [[n_C, n_gamma]]        # index for best gamma/C

                accuracity_matrix[gamma_str].append([tpr, fpr])  # dict with all acc in grid search




    ind_names = ["{:.2e}".format(C_f(i)) for i in C_range]  # names for DataFrame indexes gamma=.. podaci
    accuracity_matrix = pd.DataFrame(accuracity_matrix, index=ind_names)    # DataFrame made from dict


    # hyperparameter C are stored in row,gamma in columns
    accuracity_matrix.to_excel(Acc_grid_search, sheet_name="radial_base", startcol=0, startrow=0)
    Acc_grid_search.save()  # Saving DataFrame

    workbook = openpyxl.load_workbook(xlsx_name)        # open xlsx file
    workbook.get_sheet_by_name("radial_base").cell(row=1, column=1).value = "C/gamma"  # [0,0] cell
    workbook.get_sheet_by_name("radial_base").cell(row=1, column=1).fill = \
        openpyxl.styles.PatternFill("solid", fgColor="00FFFF00")                    # color first cell

    # Iteration for best possible hyperparameters in best_grid_acc and changing cell color
    for position in best_grid_acc:
        workbook.get_sheet_by_name("radial_base").cell(row=2 + position[0], column=2 + position[1]).fill = \
            openpyxl.styles.PatternFill("solid", fgColor="00FF0000")

    workbook.save(xlsx_name)





if __name__ == "__main__":

    Model_1(csv_data=csv_input)










